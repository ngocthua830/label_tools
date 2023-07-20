import os
import json
import shutil

import openpyxl


def sheet2json(sheet_fpath, json_fpath):
    wookbook = openpyxl.load_workbook(sheet_fpath)

    worksheet = wookbook.active

    list_of_row = []
    dict_key = []

    for col in worksheet.iter_cols(1, worksheet.max_column):
        dict_key.append(col[0].value)

    for i in range(1, worksheet.max_row):
        row_dict = {"row_index": i}
        for j, col in enumerate(worksheet.iter_cols(1, worksheet.max_column)):
            row_dict[dict_key[j]] = col[i].value
        list_of_row.append(row_dict)

    with open(json_fpath, "w") as f:
        json.dump(list_of_row, f)


def sheet2json_db(sheet_folder, json_folder):
    if not os.path.exists(json_folder):
        os.mkdir(json_folder)
    else:
        shutil.rmtree(json_folder)
        os.mkdir(json_folder)

    for sheet_fname in os.listdir(sheet_folder):
        sheet_fname_woext = sheet_fname.split(".")[0]
        sheet_fpath = os.path.join(sheet_folder, sheet_fname)

        json_fname = "%s.json" % sheet_fname_woext
        json_fpath = os.path.join(json_folder, json_fname)

        sheet2json(sheet_fpath, json_fpath)


def main():
    sheet_folder = ""
    json_folder = ""
    sheet2json_db(sheet_folder, json_folder)


if __name__ == "__main__":
    main()
